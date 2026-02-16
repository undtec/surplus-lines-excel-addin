#!/usr/bin/env python3
"""
Create Surplus Lines Tax Excel Add-in Test Workbook
Uses the unified SLAPI function matching Google Sheets integration
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_test_workbook():
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    subheader_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2E3C43", end_color="2E3C43", fill_type="solid")
    accent_fill = PatternFill(start_color="3DE2A0", end_color="3DE2A0", fill_type="solid")
    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    code_font = Font(name="Consolas", size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # =========================================================================
    # Sheet 1: Quick Reference (FIRST TAB)
    # =========================================================================
    ws_qr = wb.active
    ws_qr.title = "Quick Reference"

    # Title
    ws_qr['A1'] = "Surplus Lines Tax - Excel Function Reference"
    ws_qr['A1'].font = Font(bold=True, size=14)
    ws_qr.merge_cells('A1:D1')

    # Table header
    ws_qr['A3'] = "Function"
    ws_qr['B3'] = "Description"
    ws_qr['C3'] = "Example"
    ws_qr['D3'] = "Returns"
    for col in ['A', 'B', 'C', 'D']:
        ws_qr[f'{col}3'].font = Font(bold=True, size=11, color="FFFFFF")
        ws_qr[f'{col}3'].fill = PatternFill(start_color="2E3C43", end_color="2E3C43", fill_type="solid")
        ws_qr[f'{col}3'].alignment = Alignment(vertical='center', wrap_text=True)

    # SLAPI function entry
    ws_qr['A4'] = "SLTAX.SLAPI(calculationType, effectiveDate, stateCode, premiumAmount)"
    ws_qr['A4'].font = code_font
    ws_qr['A4'].alignment = Alignment(vertical='top', wrap_text=True)

    ws_qr['B4'] = "Unified function for tax calculations and rate lookups"
    ws_qr['B4'].alignment = Alignment(vertical='top', wrap_text=True)

    ws_qr['C4'] = 'SLTAX.SLAPI("Tax", "", "Texas", 10000)\n\nSLTAX.SLAPI("Rate", "", "Florida")\n\nSLTAX.SLAPI("Tax", "2020-01-01", "Texas", 10000)'
    ws_qr['C4'].font = code_font
    ws_qr['C4'].alignment = Alignment(vertical='top', wrap_text=True)

    ws_qr['D4'] = "Tax: 2-4 rows × 2 cols\n(Label | Value)\n\nRate: 9-11 rows × 2 cols\n(Field | Value)"
    ws_qr['D4'].alignment = Alignment(vertical='top', wrap_text=True)

    # Set row height
    ws_qr.row_dimensions[4].height = 80

    # API Key Setup section
    ws_qr['A6'] = "API Key Setup:"
    ws_qr['A6'].font = Font(bold=True, size=11)
    ws_qr['A7'] = "1. Click 'Surplus Lines Tax' in the ribbon → 'Settings'"
    ws_qr['A8'] = "2. Enter your API key from app.surpluslinesapi.com"
    ws_qr['A9'] = "3. Click 'Save API Key'"

    ws_qr['A11'] = "Get your API key: https://app.surpluslinesapi.com"
    ws_qr['A11'].font = Font(color="0563C1", underline="single")

    # Usage Notes
    ws_qr['A13'] = "Usage Notes:"
    ws_qr['A13'].font = Font(bold=True, size=11)
    ws_qr['A14'] = "• Functions that return multiple values will SPILL into adjacent cells"
    ws_qr['A15'] = "• Make sure cells to the right/below are empty for spill functions"
    ws_qr['A16'] = "• Historical queries automatically fall back to current rates if data unavailable"
    ws_qr['A17'] = "• Cost: $0.38 per query (100 free queries included)"

    # Set column widths
    ws_qr.column_dimensions['A'].width = 60
    ws_qr.column_dimensions['B'].width = 30
    ws_qr.column_dimensions['C'].width = 50
    ws_qr.column_dimensions['D'].width = 30

    # =========================================================================
    # Sheet 2: Tax Calculator
    # =========================================================================
    ws1 = wb.create_sheet("Tax Calculator")

    # Title
    ws1['A1'] = "Surplus Lines Tax Calculator"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:G1')

    # Section 1: Current Tax Calculations
    ws1['A3'] = "State"
    ws1['B3'] = "Premium"
    ws1['C3'] = "Formula"
    ws1['D3'] = "Label"
    ws1['E3'] = "Amount"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws1[f'{col}3'].font = subheader_font
        ws1[f'{col}3'].fill = accent_fill

    # Test data - Column C shows formula as text (no =), Column D has actual formula
    test_data = [
        ("Texas", 10000),
        ("California", 25000),
        ("Florida", 15000),
        ("New York", 50000),
        ("Illinois", 7500),
    ]

    row = 4
    for state, premium in test_data:
        ws1[f'A{row}'] = state
        ws1[f'B{row}'] = premium
        ws1[f'C{row}'] = f'SLTAX.SLAPI("Tax", "", A{row}, B{row})'
        ws1[f'C{row}'].font = code_font
        ws1[f'D{row}'] = f'=SLTAX.SLAPI("Tax", "", A{row}, B{row})'
        # D and E will be filled by spilling array (2 rows x 2 columns)
        row += 3  # Leave 2 rows for results + 1 buffer

    # Section 2: Historical Tax Calculation with Fallback
    historical_row = row + 2
    ws1[f'A{historical_row}'] = "Historical Tax Calculation (with Automatic Fallback)"
    ws1[f'A{historical_row}'].font = Font(bold=True, size=12)
    ws1.merge_cells(f'A{historical_row}:G{historical_row}')

    historical_row += 1
    ws1[f'A{historical_row}'] = "State"
    ws1[f'B{historical_row}'] = "Premium"
    ws1[f'C{historical_row}'] = "Date"
    ws1[f'D{historical_row}'] = "Formula"
    ws1[f'E{historical_row}'] = "Label"
    ws1[f'F{historical_row}'] = "Value"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws1[f'{col}{historical_row}'].font = subheader_font
        ws1[f'{col}{historical_row}'].fill = light_fill

    historical_row += 1
    ws1[f'A{historical_row}'] = "Texas"
    ws1[f'B{historical_row}'] = 10000
    ws1[f'C{historical_row}'] = "2020-01-01"
    ws1[f'D{historical_row}'] = f'SLTAX.SLAPI("Tax", C{historical_row}, A{historical_row}, B{historical_row})'
    ws1[f'D{historical_row}'].font = code_font
    ws1[f'E{historical_row}'] = f'=SLTAX.SLAPI("Tax", C{historical_row}, A{historical_row}, B{historical_row})'
    # Spills 2-4 rows × 2 columns (with fallback notice)

    historical_row += 5
    ws1[f'A{historical_row}'] = "Note: Historical queries may return 2-4 rows depending on data availability"
    ws1[f'A{historical_row}'].font = Font(italic=True, size=9, color="666666")
    ws1.merge_cells(f'A{historical_row}:G{historical_row}')

    # Adjust column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 45
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15

    # =========================================================================
    # Sheet 3: Rate Lookup
    # =========================================================================
    ws2 = wb.create_sheet("Rate Lookup")

    ws2['A1'] = "Tax Rate Lookup (SLAPI)"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:D1')

    # Section 1: Current Rates
    ws2['A3'] = "Current Rates"
    ws2['A3'].font = Font(bold=True, size=12)
    ws2.merge_cells('A3:D3')

    ws2['A4'] = "State"
    ws2['B4'] = "Formula"
    ws2['C4'] = "Results"
    ws2['A4'].font = subheader_font
    ws2['B4'].font = subheader_font
    ws2['C4'].font = subheader_font
    ws2['A4'].fill = accent_fill
    ws2['B4'].fill = accent_fill
    ws2['C4'].fill = accent_fill

    rate_states = ["Texas", "California", "Florida"]
    row = 5
    for state in rate_states:
        ws2[f'A{row}'] = state
        ws2[f'B{row}'] = f'SLTAX.SLAPI("Rate", "", A{row})'
        ws2[f'B{row}'].font = code_font
        ws2[f'C{row}'] = f'=SLTAX.SLAPI("Rate", "", A{row})'
        # Spills 9 rows
        row += 11  # Leave space for 9 rows + 2 buffer

    # Section 2: Historical Rates
    historical_row = row + 2
    ws2[f'A{historical_row}'] = "Historical Rates"
    ws2[f'A{historical_row}'].font = Font(bold=True, size=12)
    ws2.merge_cells(f'A{historical_row}:D{historical_row}')

    historical_row += 1
    ws2[f'A{historical_row}'] = "State"
    ws2[f'B{historical_row}'] = "Date"
    ws2[f'C{historical_row}'] = "Formula"
    ws2[f'D{historical_row}'] = "Results"
    for col in ['A', 'B', 'C', 'D']:
        ws2[f'{col}{historical_row}'].font = subheader_font
        ws2[f'{col}{historical_row}'].fill = light_fill

    historical_row += 1
    ws2[f'A{historical_row}'] = "Iowa"
    ws2[f'B{historical_row}'] = "2024-06-15"
    ws2[f'C{historical_row}'] = f'SLTAX.SLAPI("Rate", B{historical_row}, A{historical_row})'
    ws2[f'C{historical_row}'].font = code_font
    ws2[f'D{historical_row}'] = f'=SLTAX.SLAPI("Rate", B{historical_row}, A{historical_row})'
    # Spills 9 rows

    historical_row += 11  # Leave space
    ws2[f'A{historical_row}'] = "Historical with Fallback"
    ws2[f'A{historical_row}'].font = Font(bold=True, size=12)
    ws2.merge_cells(f'A{historical_row}:D{historical_row}')

    historical_row += 1
    ws2[f'A{historical_row}'] = "Texas"
    ws2[f'B{historical_row}'] = "2010-01-01"
    ws2[f'C{historical_row}'] = f'SLTAX.SLAPI("Rate", B{historical_row}, A{historical_row})'
    ws2[f'C{historical_row}'].font = code_font
    ws2[f'D{historical_row}'] = f'=SLTAX.SLAPI("Rate", B{historical_row}, A{historical_row})'
    # Spills 11 rows (includes fallback notice)

    historical_row += 12
    ws2[f'A{historical_row}'] = "Note: Rate lookups return 9 rows normally, 11 rows if fallback occurs"
    ws2[f'A{historical_row}'].font = Font(italic=True, size=9, color="666666")
    ws2.merge_cells(f'A{historical_row}:D{historical_row}')

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 55
    ws2.column_dimensions['D'].width = 20

    # =========================================================================
    # Save Workbook
    # =========================================================================
    output_path = os.path.join(os.path.dirname(__file__), "SurplusLinesTax-Test-Template.xlsx")
    wb.save(output_path)
    print(f"✅ Created test workbook: {output_path}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    create_test_workbook()
